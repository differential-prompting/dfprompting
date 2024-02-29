import json
import coverage
import requests
import os
import re
import subprocess
from openpyxl import Workbook

def run():
    # Set the proxy port, or you won't be able to connect to the api
    os.environ["http_proxy"] = "127.0.0.1:7890"  # os.environ["http_proxy"] = "http://<代理ip>:<代理端口>"
    os.environ["https_proxy"] = "127.0.0.1:7890"  # os.environ["https_proxy"] = "http://<代理ip>:<代理端口>"

    key = input("Please enter the valid key value：")
    url = f'https://generativelanguage.googleapis.com/v1/models/gemini-pro:generateContent?key={key}'
    # url = f'https://generativelanguage.googleapis.com/v1/models/gemini-pro:generateContent?key=AIzaSyADGj4bZB0nTYerFluByWCeP8cDDRxYrT8'
    headers = {'Content-Type': 'application/json'}

    # Create a new workbook
    workbook = Workbook()

    # Gets the default worksheet
    sheet = workbook.active

    # Set the header line
    sheet.cell(row=1, column=2, value="Is a Failure-inducing test case found?")
    sheet.cell(row=1, column=3, value="Test input")
    sheet.cell(row=1, column=4, value="Assertion")

    def remove_code_markers(input_text):
        # Use regular expressions to remove  "```" and "python"
        cleaned_text = re.sub(r'```|python|`', '', input_text)
        return cleaned_text

    def remove_test_input(text):
        lines = text.split('\n')
        cleaned_lines = []
        for line in lines:
            cleaned_line = line.replace('Test input:', '').strip()  # Remove the identifier and remove the trailing whitespace
            if cleaned_line:  # If the cleaned row is not empty, it is added to the list
                cleaned_lines.append(cleaned_line)
        return [cleaned_lines]  # Returns another list containing the processed text list

    def count_test_inputs(filename):
        count = 0
        with open(filename, 'r') as file:
            for line in file:
                line = line.strip()
                if line.startswith('Test input:'):
                    count += 1
        return count

    def read_test_cases(filename):
        test_cases = []

        with open(filename, 'r') as file:
            current_test_case = []
            for line in file:
                line = line.strip()
                if line:  # Checks whether the row is empty
                    if line.startswith('Test input:'):
                        if current_test_case:
                            test_cases.append(current_test_case)
                            current_test_case = []
                    else:
                        current_test_case.append(line)

            if current_test_case:
                test_cases.append(current_test_case)

        return test_cases

    def run_command_with_pipes(command, input_data):
        # Run the command and return its output as a result
        process = subprocess.Popen(command, stdin=subprocess.PIPE, stdout=subprocess.PIPE)
        return process.communicate(input_data.encode("utf-8"))[0]

    i = 1
    while i <= 1:
        # print(f"Test {i}：")
        # generating reference versions
        # Prepare user input
        j = 1
        while j <= 2:
            user_input = "Generate a python program based on the following intention.Only generate the program, not generate any explain."
            file_path_1 = "intention.txt"
            with open(file_path_1, encoding='utf-8') as file:
                file_content_1 = file.read()
            combined_text_1 = user_input + "\n" + file_content_1

            data = {
                "contents": [
                    {
                        "parts": [{"text": f"{combined_text_1}"}]
                    }
                ]
            }
            print("Waiting for Gemini...")
            response = requests.post(url, headers=headers, json=data)
            # print(f"response status_code: {response.status_code}")
            if response.status_code == 500:
                print(response.status_code)
                continue
            else:
                try:
                    reply_content = json.loads(json.dumps(response.json(), indent=4, ensure_ascii=False))
                    # print(reply_content)
                    ai_reply = reply_content['candidates'][0]['content']['parts'][0]['text']
                    print("AI:", ai_reply)
                    # Determine the directory
                    output_directory = f"test_{i}"
                    output_file_1 = f"{output_directory}/reference_version_{j}.py"

                    # Make sure the directory exists, and create it if it doesn't
                    os.makedirs(output_directory, exist_ok=True)

                    ai_reply = remove_code_markers(ai_reply)
                    # Writing the AI's response to a file
                    with open(output_file_1, 'w') as f:
                        f.write(ai_reply)

                    j += 1

                except Exception as e:
                    print(e)
                    continue

        print("Reference versions generating end!")

        coverages = []
        coverage_num = 0
        found = False
        coverage_found = False
        test_num = 1
        while test_num <= 10:
            print(f"Test inputs generation {test_num}:")
            # generating test inputs
            # Prepare user input
            user_input = "Please generate diverse test inputs for the following code.Please put Test input: as a separate line in front of each test input,no other explain is needed."
            file_path_2 = "PUT_code.py"
            with open(file_path_2, encoding='utf-8') as file:
                file_content_2 = file.read()
            combined_text_2 = user_input + "\n" + file_content_2

            data = {
                "contents": [
                    {
                        "parts": [{"text": f"{combined_text_2}"}]
                    }
                ]
            }
            print("Waiting for Gemini...")
            response = requests.post(url, headers=headers, json=data)

            if response.status_code == 500:
                print(response.status_code)
                continue
            else:
                try:
                    reply_content = json.loads(json.dumps(response.json(), indent=4, ensure_ascii=False))
                    # print(reply_content)
                    ai_reply = reply_content['candidates'][0]['content']['parts'][0]['text']
                    print("AI:", ai_reply)
                    output_file_2 = f"test_{i}/test_inputs_{test_num}.txt"
                    # Writing the AI's response to a file
                    with open(output_file_2, 'w') as f:
                        f.write(ai_reply)
                    test_num += 1
                except Exception as e:
                    print(e)
                    continue

            # Read the test case from the file
            test_cases = read_test_cases(output_file_2)
            num_test_cases = count_test_inputs(output_file_2)

            if len(test_cases) == 0:
                with open(output_file_2, encoding='utf-8') as file:
                    file_content = file.read()

                # Output after removing the Test input: identifier
                test_cases = remove_test_input(file_content)

            # print test cases
            for idx, test_case in enumerate(test_cases, start=1):
                # print(f"Test input {idx}:")
                # print(test_case)
                input_data = '\n'.join(test_case)
                # print(input_data)
                if not input_data:
                    # print(f"Test input {idx} error!")
                    continue
                else:
                    # reference_version_1 running result
                    reference_version_1_file = f"test_{i}/reference_version_1.py"
                    # executable code
                    try:
                        result_1 = subprocess.run(["python", reference_version_1_file], input=input_data,
                                                  capture_output=True, text=True, timeout=10)
                        if result_1.returncode != 0:
                            # print("reference_version_1执行出错")
                            # print("错误信息:", result_1.stderr)
                            continue
                    except subprocess.TimeoutExpired:
                        # print("reference_version_1执行超时.")
                        continue
                    result_1_output = result_1.stdout
                    # 打印输出
                    # print("Output_1:")
                    # print(result_1_output)

                    # reference_version_2运行结果
                    reference_version_2_file = f"test_{i}/reference_version_2.py"
                    try:
                        # executable code
                        result_2 = subprocess.run(["python", reference_version_2_file], input=input_data,
                                                  capture_output=True, text=True, timeout=10)
                        if result_2.returncode != 0:
                            # print("reference_version_2执行出错")
                            # print("错误信息:", result_2.stderr)
                            continue
                    except subprocess.TimeoutExpired:
                        # print("reference_version_2执行超时.")
                        continue

                    result_2_output = result_2.stdout
                    # 打印输出
                    # print("Output_2:")
                    # print(result_2_output)

                    if result_1_output != result_2_output:
                        continue
                    else:
                        # code running result
                        code_file = "PUT_code.py"
                        try:
                            # executable code
                            result_3 = subprocess.run(["python", code_file], input=input_data, capture_output=True,
                                                      text=True, timeout=10)

                        except subprocess.TimeoutExpired:
                            # print("PUT执行超时.")
                            continue

                        PUT_output = result_3.stdout
                        # 打印输出
                        # print("Output_code:")
                        # print(PUT_output)

                        if result_1_output == PUT_output:
                            # Define the command to run
                            command = ["coverage", "run", "PUT_code.py"]

                            # Execute the command and get the output
                            output = run_command_with_pipes(command, input_data)

                            # Generate a coverage report
                            cov = coverage.Coverage()
                            cov.load()  # Load the coverage data

                            # Output coverage report to a file
                            cov_report_file = "coverage_report.txt"
                            with open(cov_report_file, 'w') as report_file:
                                cov.report(file=report_file)

                            # Read the coverage report from the file
                            with open(cov_report_file, encoding='utf-8') as report_file:
                                coverage_report = report_file.read()

                            # Split coverage is reported in behavioral units
                            report_lines = coverage_report.split('\n')

                            # Initialize a variable to store coverage information of code.py
                            code_coverage = None

                            # Iterate over each line of the report
                            for line in report_lines:
                                # Look for the line that contains code.py
                                if 'code.py' in line:
                                    # Extract coverage information (assuming the coverage information is in the third column, separated by whitespace)
                                    code_coverage = line.split()[3]
                                    break
                            # print("code coverage:", code_coverage)
                            # Print coverage information for code.py
                            if code_coverage == '0':
                                code_coverage = float(code_coverage)
                            else:
                                code_coverage = float(code_coverage.strip('%'))
                            # print("code coverage:", code_coverage)

                            if coverage_num == 0:
                                coverages.append(code_coverage)
                                coverage_num += 1
                            else:
                                if code_coverage <= coverages[coverage_num - 1]:
                                    coverages.append(code_coverage)
                                    coverage_num += 1
                                    if coverage_num == 10:
                                        print("Branch coverage saturated.")
                                        coverage_found = True
                                        break
                                    else:
                                        continue
                                else:
                                    coverages = []
                                    coverage_num = 0
                                    coverages.append(code_coverage)
                                    coverage_num += 1
                                    continue

                        else:
                            row = i + 1
                            sheet.cell(row=row, column=1, value=i)
                            sheet.cell(row=row, column=2, value="Yes")
                            sheet.cell(row=row, column=3, value=input_data)
                            sheet.cell(row=row, column=4, value=result_1_output)
                            print(f"The result of test {i}:")
                            print("Failure-Inducing Test is")
                            print(input_data)
                            print()
                            print("The assert is")
                            print(result_1_output)
                            found = True
                            break

            if found:
                break

            if coverage_found:
                break

        if not found:
            row = i + 1
            sheet.cell(row=row, column=1, value=i)
            sheet.cell(row=row, column=2, value="NO")
            print(f"The result of test {i}:")
            print("No failure-inducing test is found.")

        i += 1


    file_path_coverage_1 = "coverage_report.txt"
    if os.path.exists(file_path_coverage_1):
        os.unlink(file_path_coverage_1)


    file_path_coverage_2 = ".coverage"
    if os.path.exists(file_path_coverage_2):
        os.unlink(file_path_coverage_2)

    workbook.save(filename="results.xlsx")
    print("End!")


if __name__ == '__main__':
    run()



